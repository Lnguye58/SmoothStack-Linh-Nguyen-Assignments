from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import logging
import os 
import webbrowser as wb


logging.basicConfig(filename= 'excel_month_report.log', filemode='w', 
                    format='%(asctime)s:%(name)s:%(levelname)s - %(message)s', level=logging.INFO)

def Check_Excel():
    folder = os.listdir('.')
    excel_file = [files for files in folder if '.xlsx' in files]
    get_files = False
    exit_program = False
    
    while not get_files:
        try:
            for files in enumerate(excel_file, 1):
                print(files)
                #raise error here when no file in folder
            selected = input("Select and type the number to open the file or type 'Exit' to exit program: ")
            if selected.isdigit():
                if int(selected) <= (len(excel_file)):
                    get_files = True
                else:
                    print('Invalid number selected, try select the current availible number presented')
                    logging.warning('Invalid number selected, try select the current availible number presented')
                    wb.open('excel_month_report.log')
                    continue
            if selected == "Exit":
                    print("Exited Program")
                    exit_program = True
            else:
                continue

        except Exception as e:
            print(e)

        finally:
            if exit_program == True:
                quit()
            if get_files:
                pass
            else:
                print('Unable to find an excel workbook in the folder, check if workbook in folder')
                logging.warning('Unable to find an excel workbook in the folder, check if workbook in folder')
                Check_Excel()
                #raise FileNotFoundError
                
    return excel_file[int(selected)-1][:-5]

def Excel():
    #set the chosen excel file and load the excel workbook
    excel_file = Check_Excel()
    excel_wb = load_workbook(excel_file + '.xlsx')

    #Convert letters to digits for excel to read
    months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}
    years = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']

    #String slicing the excel to get the month and year
    year_month = [date for date in excel_file.split("_")[::-1] if date.capitalize() in months or date in years]
    #format it as year-month, 2018-01 // 2018-03
    y_m = (f'{year_month[0]}-{months[year_month[1].capitalize()]}')
    
    #Get data from 'Summary Rolling MoM' worksheet
    ws0 = excel_wb['Summary Rolling MoM']
    
    #find the row of the specific month's and year's
    row = None
    for cell in ws0['A']:
        if y_m in str(cell.value):
            row = cell.row
            #print("Current row number of selected month is: " + str(row))

    #loop through current row of month
    row_data = [cell for cell in ws0.iter_rows(min_row = row, max_row = row, values_only = True)]
    #print(row_data)

    #Get data in row that doesnt contain None cell
    ws0_data = [item for item in row_data[0][1:] if item != None]
    #print(ws0_data)
    
    #Access 2nd sheet
    ws1 = excel_wb["VOC Rolling MoM"]
    col_number = 0
    month_not_found = False
    year_not_found = False
    
    #Check if both month-year in column 
    for dates in ws1['1']:
        col_number += 1
        if y_m in str(dates.value):
            #print(dates)
            year_not_found = True
            break
        else:
            continue
    #If there's no year find the month
    if not year_not_found:
            col_number = 0
            for dates in ws1['1']:
                col_number += 1
                if str(dates.value) in months:  
                    year_not_found = True        
                    break
                else:
                    continue
    #If year and month not found, raise file error
    if not year_not_found and not month_not_found:
            col_number = 0
            for dates in ws1['1']:
                col_number += 1
                if str(dates.value) not in months and str(dates.value) not in years:   
                    print("Unable to locate data based on month and year input, check if input is correct.")
                    logging.warning("Unable to locate data based on month and year input, check if input is correct.")
                    wb.open('excel_month_report.log')
                    raise FileNotFoundError
                else:
                    continue
    
    #Get data when found targeted month/year
    ws1_data = [data.value for data in ws1[get_column_letter(col_number)][1:] if data.value != None and int(data.value) != 0]
    
    return Log_File(year_month, ws0_data, ws1_data)

def Log_File(year_month, ws0_data, ws1_data):

    #logging for Summary  VOC
    logging.info(f"Month of {year_month[1].capitalize()}, {year_month[0]} ")
    logging.info(f"Calls Offered: {ws0_data[0]}")
    logging.info(f"Abandon after 30s: {ws0_data[1]*100}%")
    logging.info(f"FCR : {ws0_data[2]*100}0%")
    logging.info(f"DSAT : {ws0_data[3]*100}0%")
    logging.info(f"CSAT : {ws0_data[4]*100}0%")

    #logging for VOC
    logging.info(f"In Net Promoter Score - Base Size = {ws1_data[0]}")
    if ws1_data[1] >= 200:
        logging.info(f"Promoters: {ws1_data[1]}, good")
    else:
        logging.info(f"Promoters: {ws1_data[1]}, bad")
    if ws1_data[2] >= 100:
        logging.info(f"Passives: {ws1_data[2]}, good")
    else:
        logging.info(f"Passives: {ws1_data[2]}, bad")
    if ws1_data[3] >= 100:
        logging.info(f"Decractors: {ws1_data[3]}, good")
    else:
        logging.info(f"Decractors: {ws1_data[3]}, bad")
    Logging()

def Logging():
    u_input = input("Open Log file for Month's Report? (Yes to Open // No to Exit Program)  ")
    exiting = False

    try:
        if u_input== "Yes" or u_input == "YES":
            wb.open('excel_month_report.log')
            exiting = True
        if u_input == "No" or u_input == "NO":
            print("Program Exiting")
            exiting = True

    except Exception as e:
        print(e)

    finally:
        if exiting:
            print("Program Exit")
            quit()
        else:
            print("Invalid Answer, Type Yes or No ")
            Logging()

if __name__ == "__main__":
    Excel()

